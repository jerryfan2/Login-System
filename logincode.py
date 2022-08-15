import openpyxl
from openpyxl.chart import Reference, BarChart, Series
import logging
import os


if os.path.exists('C:/Users/jerry/Desktop/Work/School/Grade 12 - \
2020-2021/Capstone Project/PyScripts/Login System/LoginSystem.xlsx'):
    wb = openpyxl.load_workbook("LoginSystem.xlsx")
else:
    wb = openpyxl.Workbook()

wb.active.title = "User Information"
ws = wb["User Information"]

ws.cell(row=1, column=1, value="Username")
ws["b1"] = "Email"
ws["c1"] = "Password"
ws["e1"] = "Times Logged On"

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

formatter = logging.Formatter("%(asctime)s:%(levelname)s:%(name)s:%(message)s")

file_handler = logging.FileHandler("loginsystemlog.log")
file_handler.setFormatter(formatter)

logger.addHandler(file_handler)


def signup():
    username = None
    email = None
    password = None

    def create_username():
        nonlocal username
        while True:
            username = input("Enter username: ")
            for value in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                if username in value:
                    print("Username already taken")
                    break
            else:
                ws[f'a{len(ws["a"])+1}'] = username
                break

    def create_email():
        nonlocal email
        while True:
            email = input("Enter email: ")
            for value in ws.iter_rows(min_col=2, max_col=2, values_only=True):
                if email in value:
                    print("Email already in use")
                    break
            else:
                for cell in ws["b"]:
                    if cell.value is None:
                        cell.value = email
                break

    def create_password():
        nonlocal password
        while True:
            password1 = input("Enter password: ")
            password2 = input("Re-enter password: ")
            if password1 != password2:
                print("Passwords do not match")
            else:
                password = password1
                break
        for cell in ws["c"]:
            if cell.value is None:
                cell.value = password

    def welcome_msg():
        print(f"Welcome, {username}!")

    def log():
        logger.info(f"{username} has signed up")

    create_username()
    create_email()
    create_password()
    welcome_msg()
    log()
    for cell in ws["e"]:
        if cell.value is None:
            cell.value = 1
            break


def login():
    info_row = None
    username = None
    password = None

    def get_username():
        nonlocal info_row
        nonlocal username
        while True:
            username = input("Enter username: ")
            for cell in ws["a"]:
                if cell.value == username:
                    info_row = cell.row
                    break
            else:
                print("Invalid username")
                continue
            break

    def get_password():
        nonlocal password
        while True:
            password = input("Enter password: ")
            if password == ws[f"c{info_row}"].value:
                break
            else:
                print("Incorrect password")

    def welcome_back_msg():
        print(f"Welcome back, {username}!")

    def log():
        logger.info(f"{username} has logged in")

    get_username()
    get_password()
    welcome_back_msg()
    log()
    ws[f"e{info_row}"].value += 1


print("Welcome to Jerry's log in system!")

while True:
    choose = input("(L)og in or (S)ign up: ").lower()
    if choose == "s" or choose == "l":
        break
    else:
        print("Please enter a valid choice")

if choose == "s":
    signup()

if choose == "l":
    login()

chart = BarChart()
chart.title = "Times Logged On per User"
chart.y_axis.title = "Times Logged On"
chart.x_axis.title = "User"
chart.y_axis.majorUnit = 1

values = Reference(ws, min_col=5, max_col=5, min_row=2, max_row=ws.max_row)

series = Series(values, title="User")
chart.series.append(series)

titles = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)
chart.set_categories(titles)

ws.add_chart(chart, "g2")

wb.save("LoginSystem.xlsx")
